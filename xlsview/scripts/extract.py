# -*- coding: utf-8 -*-
"""XLSX → JSON（供 xlsview.nvim 渲染）。

依赖: openpyxl

用法:
  python -X utf8 extract.py <xlsx_path> <out_json> [max_rows] [max_cols]

stdout: OK <sheet_count> 或 ERROR <msg>
"""
from __future__ import annotations

import json
import os
import sys
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl required (pip install openpyxl)", file=sys.stderr)
    sys.exit(2)


def color_to_hex(color) -> Optional[str]:
    """openpyxl Color → #rrggbb；主题色/自动返回 None。"""
    if color is None:
        return None
    # type: rgb / theme / indexed / auto
    rgb = getattr(color, "rgb", None)
    if rgb is None and isinstance(color, str):
        rgb = color
    if not rgb:
        return None
    s = str(rgb).strip()
    if s.lower() in ("00000000", "none", "auto"):
        return None
    # AARRGGBB or RRGGBB
    s = s.lstrip("#")
    if len(s) == 8:
        s = s[2:]  # drop alpha
    if len(s) != 6:
        return None
    try:
        int(s, 16)
    except ValueError:
        return None
    # 纯白底可保留；全透明已滤
    return "#" + s.lower()


def cell_style(cell) -> Dict[str, Any]:
    bold = italic = mono = False
    fg = None
    bg = None
    size = 11.0
    font = cell.font
    if font is not None:
        bold = bool(font.bold)
        italic = bool(font.italic)
        if font.name:
            n = font.name.lower()
            if any(k in n for k in ("consolas", "courier", "mono", "menlo", "cascadia")):
                mono = True
        if font.size:
            try:
                size = float(font.size)
            except (TypeError, ValueError):
                pass
        fg = color_to_hex(font.color)
    fill = cell.fill
    if fill is not None and getattr(fill, "fill_type", None) not in (None, "none"):
        # solid / pattern
        fg_color = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
        bg = color_to_hex(fg_color)
    align = None
    if cell.alignment and cell.alignment.horizontal:
        h = str(cell.alignment.horizontal).lower()
        if h in ("left", "right", "center", "general"):
            align = "left" if h == "general" else h
    return {
        "bold": bold,
        "italic": italic,
        "mono": mono,
        "color": fg or "#000000",
        "bg": bg,
        "size": size,
        "align": align or "left",
    }


def cell_value(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        # 避免 1.0 冗长；保留有意义小数
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        return ("%g" % v)
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d %H:%M:%S").rstrip(" 00:00:00").rstrip()
        except Exception:
            return str(v)
    # 富文本
    if hasattr(v, "plain"):
        try:
            return str(v.plain)
        except Exception:
            pass
    s = str(v)
    # 公式结果：data_only 时已是值；否则显示公式
    return s.replace("\r\n", "\n").replace("\r", "\n")


def sheet_used_bounds(ws, max_rows: int, max_cols: int) -> Tuple[int, int, int, int]:
    """返回 1-based inclusive (min_r, min_c, max_r, max_c)。"""
    if ws.calculate_dimension() == "A1:A1" and ws["A1"].value is None:
        return 1, 1, 1, 1
    # openpyxl dimensions 可能偏大（格式刷过的空区）
    min_r, min_c = ws.min_row or 1, ws.min_column or 1
    max_r, max_c = ws.max_row or 1, ws.max_column or 1
    # 收缩全空边
    def row_empty(r: int) -> bool:
        for c in range(min_c, max_c + 1):
            if ws.cell(r, c).value is not None:
                return False
        return True

    def col_empty(c: int) -> bool:
        for r in range(min_r, max_r + 1):
            if ws.cell(r, c).value is not None:
                return False
        return True

    while max_r > min_r and row_empty(max_r):
        max_r -= 1
    while min_r < max_r and row_empty(min_r):
        min_r += 1
    while max_c > min_c and col_empty(max_c):
        max_c -= 1
    while min_c < max_c and col_empty(min_c):
        min_c += 1

    # 限制
    if max_rows > 0:
        max_r = min(max_r, min_r + max_rows - 1)
    if max_cols > 0:
        max_c = min(max_c, min_c + max_cols - 1)
    return min_r, min_c, max_r, max_c


def extract_sheet(ws, max_rows: int, max_cols: int) -> Dict[str, Any]:
    min_r, min_c, max_r, max_c = sheet_used_bounds(ws, max_rows, max_cols)
    rows: List[List[Dict[str, Any]]] = []
    for r in range(min_r, max_r + 1):
        row_cells = []
        for c in range(min_c, max_c + 1):
            cell = ws.cell(r, c)
            st = cell_style(cell)
            text = cell_value(cell)
            # 多行单元格 → 单行显示（表内）
            text = text.replace("\n", " ").replace("\t", " ")
            row_cells.append(
                {
                    "text": text,
                    "bold": st["bold"],
                    "italic": st["italic"],
                    "mono": st["mono"],
                    "color": st["color"],
                    "bg": st["bg"],
                    "align": st["align"],
                    "size": st["size"],
                }
            )
        rows.append(row_cells)

    # 合并单元格信息（可选，渲染时主格保留值）
    merges = []
    try:
        for mr in ws.merged_cells.ranges:
            merges.append(str(mr))
    except Exception:
        pass

    return {
        "name": ws.title,
        "min_row": min_r,
        "min_col": min_c,
        "max_row": max_r,
        "max_col": max_c,
        "nrows": max_r - min_r + 1,
        "ncols": max_c - min_c + 1,
        "rows": rows,
        "merges": merges,
        "hidden": bool(getattr(ws, "sheet_state", "visible") != "visible"),
    }


def main() -> int:
    if len(sys.argv) < 3:
        print(
            "usage: extract.py <xlsx> <out_json> [max_rows] [max_cols]",
            file=sys.stderr,
        )
        return 2
    path = os.path.abspath(sys.argv[1])
    out_json = os.path.abspath(sys.argv[2])
    max_rows = int(sys.argv[3]) if len(sys.argv) > 3 else 500
    max_cols = int(sys.argv[4]) if len(sys.argv) > 4 else 64

    if not os.path.isfile(path):
        print(f"ERROR: file not found: {path}", file=sys.stderr)
        return 1

    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        print(f"ERROR: unsupported extension {ext} (need xlsx/xlsm)", file=sys.stderr)
        return 1

    try:
        # data_only=True 尽量显示公式结果；失败则回退
        try:
            wb = load_workbook(path, data_only=True, read_only=False)
        except Exception:
            wb = load_workbook(path, data_only=False, read_only=False)
    except Exception as e:
        print(f"ERROR: open failed: {e}", file=sys.stderr)
        return 1

    try:
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            try:
                sheets.append(extract_sheet(ws, max_rows, max_cols))
            except Exception as e:
                sheets.append(
                    {
                        "name": name,
                        "error": str(e),
                        "rows": [],
                        "nrows": 0,
                        "ncols": 0,
                    }
                )
        try:
            st = os.stat(path)
            mtime, size = int(st.st_mtime), int(st.st_size)
        except OSError:
            mtime, size = 0, 0

        props = {}
        try:
            p = wb.properties
            props = {
                "title": getattr(p, "title", None) or "",
                "creator": getattr(p, "creator", None) or "",
            }
        except Exception:
            pass

        payload = {
            "version": 1,
            "kind": "xlsx",
            "path": path,
            "mtime": mtime,
            "size": size,
            "meta": props,
            "sheet_count": len(sheets),
            "sheets": sheets,
            "max_rows": max_rows,
            "max_cols": max_cols,
        }
        os.makedirs(os.path.dirname(out_json) or ".", exist_ok=True)
        with open(out_json, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        print(f"OK {len(sheets)}")
        return 0
    finally:
        wb.close()


if __name__ == "__main__":
    sys.exit(main())
